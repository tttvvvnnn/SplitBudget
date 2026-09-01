"""Экспорт истории трат и баланса в Excel (.xlsx)."""
from __future__ import annotations

import datetime as dt
import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from app.api.dependencies import ChatContext, get_chat_context
from app.shared.balance import compute_net_balances, simplify_debts
from app.shared.models import Expense, ExpenseShare, Member

router = APIRouter(tags=["export"])


@router.get("/chats/{chat_id}/export")
async def export_xlsx(ctx: ChatContext = Depends(get_chat_context), month: str | None = None):
    query = select(Expense).where(Expense.chat_id == ctx.chat.id)
    if month:
        try:
            year, mon = (int(p) for p in month.split("-"))
            start = dt.date(year, mon, 1)
            end = dt.date(year + (mon // 12), (mon % 12) + 1, 1)
        except (ValueError, IndexError) as exc:
            raise HTTPException(status_code=400, detail="month должен быть в формате YYYY-MM") from exc
        query = query.where(Expense.expense_date >= start, Expense.expense_date < end)
    query = query.order_by(Expense.expense_date)

    result = await ctx.session.execute(query)
    expenses = result.scalars().unique().all()

    members_result = await ctx.session.execute(select(Member).where(Member.chat_id == ctx.chat.id))
    members_by_id = {m.id: m for m in members_result.scalars().all()}

    ids = [e.id for e in expenses]
    shares_by_expense: dict[int, list[ExpenseShare]] = {}
    if ids:
        shares_result = await ctx.session.execute(
            select(ExpenseShare).where(ExpenseShare.expense_id.in_(ids))
        )
        for s in shares_result.scalars().all():
            shares_by_expense.setdefault(s.expense_id, []).append(s)

    def member_label(mid: int) -> str:
        m = members_by_id.get(mid)
        if not m:
            return str(mid)
        return f"@{m.username}" if m.username else m.full_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Траты"
    headers = ["Дата", "Название", "Категория", "Сумма", "Валюта", "Оплатил", "Тип деления", "Участники и доли", "Добавил"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for e in expenses:
        shares = shares_by_expense.get(e.id, [])
        shares_text = "; ".join(f"{member_label(s.member_id)}: {s.amount}" for s in shares)
        ws.append(
            [
                e.expense_date.isoformat(),
                e.title,
                e.category,
                float(e.amount),
                ctx.chat.currency,
                member_label(e.payer_member_id),
                "поровну" if e.split_type == "equal" else "вручную",
                shares_text,
                member_label(e.created_by_member_id),
            ]
        )
    for i, width in enumerate([12, 28, 16, 12, 8, 18, 12, 40, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    balances = await compute_net_balances(ctx.session, ctx.chat.id)
    debts = simplify_debts(balances)

    ws2 = wb.create_sheet("Баланс")
    ws2.append(["Участник", "Баланс (+ должны ему, − должен он)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for mid, net in balances.items():
        ws2.append([member_label(mid), float(net)])

    ws2.append([])
    ws2.append(["Кто кому должен"])
    ws2[ws2.max_row][0].font = Font(bold=True)
    for frm, to, amount in debts:
        ws2.append([f"{member_label(frm)} → {member_label(to)}", float(amount)])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"expenses_{ctx.chat.id}" + (f"_{month}" if month else "") + ".xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
